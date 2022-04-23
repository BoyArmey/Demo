import openpyxl
from aip import AipOcr

'''
  身份证信息识别并写入Excel文件
'''


""" 你的 APPID AK SK """
APP_ID = ''
API_KEY = ''
SECRET_KEY = ''
client = AipOcr(APP_ID, API_KEY, SECRET_KEY)


# 读取文件
def get_file_content(filePath):
    with open(filePath, 'rb') as fp:
        return fp.read()

def OCR():
    #身份证照片路径
    fileName1 = r'C:\Users\...\Desktop\data\1.jpg'     # 图片路径

    image = get_file_content(fileName1)
    # 识别身份证正面
    idCardSide = "front"
    # 识别身份证背面
    # idCardSide = "back"
    # 调用身份证识别
    client.idcard(image, idCardSide);
    # 如果有可选参数
    options = {}
    options["detect_direction"] = "true"
    # 是否检测图像朝向，默认不检测
    options["detect_risk"] = "false"
    """ 带参数调用身份证识别 """
    result = client.idcard(image, idCardSide, options)
    # 获取返回识别结果
    if isinstance(result, dict):
        words = result['words_result']
        lst = list()
        lst.append(words['姓名']['words'])
        lst.append(words['公民身份号码']['words'])
        lst.append(words['性别']['words'])
        lst.append(words['民族']['words'])
        lst.append(words['住址']['words'])
        print(lst)
        return lst

def writeExcel(lst):
    wb = openpyxl.load_workbook(r'D:\Example\pythonCode\...\data.xlsx')  # Excel文件路径
    sheet = wb.worksheets[0]     # 选择工作表
    rowNum = sheet.max_row       # 获取工作表的最大行数
    # 写入数据
    for i in range(1,6):
        sheet.cell(rowNum+1,i).value = lst[i - 1]
    wb.save('data.xlsx')

if __name__ == '__main__':
    lst = OCR()
    writeExcel(lst)
    print('OK')



